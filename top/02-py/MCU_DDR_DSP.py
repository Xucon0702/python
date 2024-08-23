
import os
import csv

import openpyxl
from openpyxl import Workbook
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
# import cairosvg
from io import BytesIO
def list_files_in_folder(folder_path):
    files = []
    for file_name in os.listdir(folder_path):
        # print(file_name)
        file_path = os.path.join(folder_path, file_name)
        # print(file_path)
        if os.path.isfile(file_path):
            files.append(file_path)
    return files

def list_dir_in_folder(folder_path):
    files = []
    for file_name in os.listdir(folder_path):
        # print(file_name)
        file_path = os.path.join(folder_path, file_name)
        # print(file_path)
        if os.path.isdir(file_path):
            files.append(file_path)
    return files


def read_mcu_dsp_ddr(in_file):
    in_f = open(in_file,'r',encoding='utf-8')
    # out_f = open(str(in_file).split('.txt')[0] + '.csv','w',encoding='utf-8')
    count = 1

    wb = Workbook()
    ws = wb.active
    # with open()
    # with open(str(in_file).split('.txt')[0] + '.csv','w',newline='') as file:
    #     writer = csv.writer(file)
    for line in in_f.readlines():
        if line[:4] == 'CPU:':
            # print(line)
            new_line = line.split()
            # print(new_line)
            ws['A' + str(count)] = count % 6
            ws['B' + str(count)] = str(new_line[1][:6])
            ws['C' + str(count)] = str( new_line[5] + new_line[6]).rstrip('%')
            count += 1
    in_f.close()
    in_f = open(in_file, 'r', encoding='utf-8')

    count = 1
    for line in in_f.readlines():
        if line[:4] == 'DDR:':
            new_line = line.split()
            # print(new_line)
            ws['E' + str(count)] = count % 3
            ws['F' + str(count)] = str(new_line[1])
            ws['G' + str(count)] = str(new_line[5])
            ws['H' + str(count)] = str(new_line[9])
            count += 1
    in_f.close()
    wb.save(str(in_file).split('.txt')[0] + '.xlsx')
    draw(str(in_file).split('.txt')[0] + '.xlsx')
def draw(in_file):
    wb =  openpyxl.load_workbook(in_file)
    ws = wb.active
    i = 1
    mpu10_y_list= []
    mcu20_y_list= []
    mcu21_y_list= []
    c6x1_y_list= []
    c6x2_y_list= []
    c7x1_y_list= []

    # x_list = [(x+ 1) for x in range(int(ws.max_row/6))]
    while i <= ws.max_row:
        mpu10_y_list.append(float(ws['C' + str(i)].value))
        mcu20_y_list.append(float(ws['C' + str(i+1)].value))
        mcu21_y_list.append(float(ws['C' + str(i+2)].value))
        c6x1_y_list.append(float(ws['C' + str(i+3)].value))
        c6x2_y_list.append(float(ws['C' + str(i+4)].value))
        c7x1_y_list.append(float(ws['C' + str(i+5)].value))
        i += 6
    i = 1
    while i <= ws.max_row:
        value = ws['E' + str(i)].value
        # print(str(value) + ' ' + str(i))
        if value is None:
            break
        i += 1
    # print(i - 1)
    read_list = []
    write_list = []
    total_list = []
    n = i - 1
    k = 1
    # print( k ,n)
    while k <= n:
        read_list.append(int(ws['G' + str(k)].value))
        write_list.append(int(ws['G' + str(k+1)].value))
        total_list.append(int(ws['G' + str(k+2)].value))
        k += 3
        # print(k ,n)
        # break
    # print(read_list)
    # print(x_list)
    # print(type(mpu10_y_list[0]))
    # x_p = np.array(x_list)
    y1_p = np.array(mpu10_y_list)
    y2_p = np.array(mcu20_y_list)
    y3_p = np.array(mcu21_y_list)
    y4_p = np.array(c6x1_y_list)
    y5_p = np.array(c6x2_y_list)
    y6_p = np.array(c7x1_y_list)


    plt.subplot(2,1,1)
    plt.title(str(in_file) + '\nMCU DSP')
    plt.plot(y1_p,label='mpu1_0')
    plt.plot(y2_p,label='mcu2_0')
    plt.plot(y3_p,label='mcu2_1')
    plt.plot(y4_p,label='c6x_1')
    plt.plot(y5_p,label='c6x_2')
    plt.plot(y6_p,label='c7x_1')
    plt.ylabel('MCU DSP loading(%)')
    plt.xlabel('Time')
    plt.yticks(range(0, 100, 20))
    # plt.ylim(0,100)
    # plt.axis([0,len(x_list),0,100])
    plt.grid(True)
    plt.legend()
    plt.subplot(2, 1, 2)
    # plt.plot([2,3,1])
    plt.title('DDR')
    plt.plot(read_list,label='read')
    plt.plot(write_list,label='write')
    plt.plot(total_list,label='total')
    plt.ylabel('DDR Rate(MB/S)')
    plt.xlabel('Time')
    plt.legend()
    plt.yticks(range(0, 10000, 1000))
    plt.show()

#两种功能1.出图  2.出表格
if __name__ == '__main__':
    # 出图
    file_path = list_dir_in_folder('./')
    for k in file_path:
        for j in list_files_in_folder(k):
            if 'mcu.txt' in str(j):
                read_mcu_dsp_ddr(j)
                print(j)


    # #出表格
    # file_path = list_files_in_folder('C:\\Users\\liuchenchen\\Desktop\\data1103\\2023-11-03_18_50_06data')
    # wb = Workbook()
    # for k in file_path:
    #     if 'svg' in str(k):
    #         print(k)
    #         s = k.split('\\')[-1]
    #         # print(s)
    #         ws = wb.create_sheet(s.split('.')[0])
    #         # png = cairosvg.svg2png(k)
    #         # image_stream = BytesIO(png)
    #         # image = Image(image_stream)
    #         # ws.add_image(image,'A1')
    #
    #         wb.save(k.split('\\')[-2] + '.xlsx')